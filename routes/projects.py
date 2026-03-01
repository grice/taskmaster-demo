from io import BytesIO
from flask import Blueprint, render_template, request, redirect, url_for, jsonify, send_file
from models import db, Project, Task, StatusUpdate, Milestone
from datetime import date

bp = Blueprint('projects', __name__)


@bp.route('/projects')
def list_projects():
    status_filter = request.args.get('status', '')
    q = Project.query
    if status_filter:
        q = q.filter_by(status=status_filter)
    projects = q.order_by(Project.start_date.desc()).all()
    return render_template('projects/list.html', projects=projects, status_filter=status_filter)


@bp.route('/projects/new', methods=['GET', 'POST'])
def new_project():
    if request.method == 'POST':
        from datetime import date
        project = Project(
            name=request.form['name'],
            description=request.form.get('description', ''),
            start_date=date.fromisoformat(request.form['start_date']) if request.form.get('start_date') else None,
            end_date=date.fromisoformat(request.form['end_date']) if request.form.get('end_date') else None,
            status=request.form.get('status', 'active'),
        )
        db.session.add(project)
        db.session.commit()
        return redirect(url_for('projects.detail', id=project.id))
    return render_template('projects/form.html', project=None)


@bp.route('/projects/<int:id>')
def detail(id):
    project = Project.query.get_or_404(id)
    # Collect all status updates across this project's tasks, newest first
    task_ids = [t.id for t in project.tasks]
    all_updates = StatusUpdate.query.filter(
        StatusUpdate.task_id.in_(task_ids)
    ).order_by(StatusUpdate.created_at.desc()).all() if task_ids else []
    upcoming_milestones = Milestone.query.filter(
        Milestone.task_id.in_(task_ids),
        Milestone.date >= date.today()
    ).order_by(Milestone.date).all() if task_ids else []
    return render_template('projects/detail.html', project=project,
                           all_updates=all_updates,
                           upcoming_milestones=upcoming_milestones)


@bp.route('/projects/<int:id>/edit', methods=['GET', 'POST'])
def edit_project(id):
    project = Project.query.get_or_404(id)
    if request.method == 'POST':
        from datetime import date
        project.name = request.form['name']
        project.description = request.form.get('description', '')
        project.start_date = date.fromisoformat(request.form['start_date']) if request.form.get('start_date') else None
        project.end_date = date.fromisoformat(request.form['end_date']) if request.form.get('end_date') else None
        project.status = request.form.get('status', 'active')
        db.session.commit()
        return redirect(url_for('projects.detail', id=project.id))
    return render_template('projects/form.html', project=project)


@bp.route('/projects/<int:id>/delete', methods=['GET', 'POST'])
def delete_project(id):
    project = Project.query.get_or_404(id)
    if request.method == 'POST':
        db.session.delete(project)
        db.session.commit()
        return redirect(url_for('projects.list_projects'))
    return render_template('confirm_delete.html',
                           title='Delete Project',
                           message=f'Are you sure you want to delete project "{project.name}" and all its tasks? This cannot be undone.',
                           cancel_url=url_for('projects.detail', id=project.id))


@bp.route('/projects/<int:id>/export/excel')
def export_excel(id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.styles.fills import FILL_SOLID
    from openpyxl.utils import get_column_letter

    project = Project.query.get_or_404(id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Gantt Data'

    # Determine max milestones across all tasks so we can size the header row
    max_ms = max((len(t.milestones) for t in project.tasks), default=0)

    # ── Header row ────────────────────────────────────────────────────────────
    fixed_headers = ['Task', 'Start Date', 'End Date', 'Status', 'Priority', 'Assignees', 'Tags']
    ms_headers = []
    for i in range(1, max_ms + 1):
        ms_headers += [f'Milestone {i}', f'Milestone {i} Date']
    headers = fixed_headers + ms_headers

    header_fill = PatternFill(fill_type=FILL_SOLID, fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    ws.row_dimensions[1].height = 20

    # ── Task rows ─────────────────────────────────────────────────────────────
    date_fmt = 'YYYY-MM-DD'
    row_fill_even = PatternFill(fill_type=FILL_SOLID, fgColor='EBF3FB')

    for row_idx, task in enumerate(project.tasks, 2):
        fill = row_fill_even if row_idx % 2 == 0 else None
        assignees = ', '.join(
            (a.person.name + (' (lead)' if a.is_lead else ''))
            for a in task.assignments
        ) or 'Unassigned'
        tags = ', '.join(t.name for t in task.tags)

        values = [
            task.title,
            task.start_date,
            task.end_date,
            task.status.replace('_', ' ').title(),
            task.priority.title(),
            assignees,
            tags,
        ]
        for ms in task.milestones:
            values += [ms.name, ms.date]
        # Pad to full width
        values += [None] * (len(headers) - len(values))

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if fill:
                cell.fill = fill
            # Format date columns
            if isinstance(value, date):
                cell.number_format = date_fmt

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [40, 14, 14, 14, 12, 30, 20] + [24, 14] * max_ms
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = 'B2'

    # ── Project metadata sheet ────────────────────────────────────────────────
    meta = wb.create_sheet('Project Info')
    meta_rows = [
        ('Project', project.name),
        ('Status', project.status.replace('_', ' ').title()),
        ('Start Date', project.start_date),
        ('End Date', project.end_date),
        ('Exported', date.today()),
        ('Tasks', len(project.tasks)),
    ]
    for r, (label, value) in enumerate(meta_rows, 1):
        meta.cell(row=r, column=1, value=label).font = Font(bold=True)
        cell = meta.cell(row=r, column=2, value=value)
        if isinstance(value, date):
            cell.number_format = date_fmt
    meta.column_dimensions['A'].width = 14
    meta.column_dimensions['B'].width = 30

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = project.name.replace(' ', '_') + '_gantt.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


@bp.route('/projects/<int:id>/gantt-data')
def gantt_data(id):
    project = Project.query.get_or_404(id)
    tasks = []
    for task in project.tasks:
        dep_ids = ','.join(f'task-{d.id}' for d in task.dependencies)
        lead = task.lead
        assignee_names = []
        for a in task.assignments:
            name = a.person.name
            if a.is_lead:
                name += ' (lead)'
            assignee_names.append(name)
        milestones = [{'name': ms.name, 'date': ms.date.isoformat(),
                       'status': ms.computed_status} for ms in task.milestones]
        tasks.append({
            'id': f'task-{task.id}',
            'name': task.title,
            'start': task.start_date.isoformat(),
            'end': task.end_date.isoformat(),
            'progress': task.progress,
            'dependencies': dep_ids,
            'custom_class': f'status-{task.status} priority-{task.priority}',
            'assignees': ', '.join(assignee_names) if assignee_names else 'Unassigned',
            'milestones': milestones,
        })
    return jsonify(tasks)
