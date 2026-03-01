"""Tests for project CRUD routes."""
from datetime import date
from io import BytesIO
import openpyxl
from models import Project
from tests.conftest import make_project, make_task, make_milestone


class TestProjectList:
    def test_empty_list(self, client):
        r = client.get('/projects')
        assert r.status_code == 200
        assert b'Projects' in r.data

    def test_shows_projects(self, client, db):
        make_project('Alpha')
        make_project('Beta')
        r = client.get('/projects')
        assert b'Alpha' in r.data
        assert b'Beta' in r.data

    def test_status_filter_active(self, client, db):
        make_project('Active One', status='active')
        make_project('Done One', status='completed')
        r = client.get('/projects?status=active')
        assert b'Active One' in r.data
        assert b'Done One' not in r.data

    def test_status_filter_heading(self, client, db):
        make_project('Active One', status='active')
        r = client.get('/projects?status=active')
        assert b'Active Projects' in r.data

    def test_no_filter_shows_all(self, client, db):
        make_project('Active One', status='active')
        make_project('Done One', status='completed')
        r = client.get('/projects')
        assert b'Active One' in r.data
        assert b'Done One' in r.data


class TestNewProject:
    def test_get_form(self, client):
        r = client.get('/projects/new')
        assert r.status_code == 200
        assert b'Project Name' in r.data

    def test_create_project(self, client, db):
        r = client.post('/projects/new', data={
            'name': 'My New Project',
            'status': 'active',
            'start_date': '2025-01-01',
            'end_date': '2025-12-31',
            'description': '',
        }, follow_redirects=True)
        assert r.status_code == 200
        project = Project.query.filter_by(name='My New Project').first()
        assert project is not None
        assert project.status == 'active'

    def test_create_redirects_to_detail(self, client, db):
        r = client.post('/projects/new', data={
            'name': 'Redirect Test',
            'status': 'active',
            'start_date': '2025-01-01',
            'end_date': '2025-12-31',
        })
        assert r.status_code == 302
        assert '/projects/' in r.headers['Location']


class TestProjectDetail:
    def test_shows_project(self, client, db):
        p = make_project('Detail Project')
        r = client.get(f'/projects/{p.id}')
        assert r.status_code == 200
        assert b'Detail Project' in r.data

    def test_shows_tasks(self, client, db):
        p = make_project()
        make_task(p, 'My Task')
        r = client.get(f'/projects/{p.id}')
        assert b'My Task' in r.data

    def test_404_for_missing(self, client):
        r = client.get('/projects/99999')
        assert r.status_code == 404

    def test_gantt_data_json(self, client, db):
        p = make_project()
        make_task(p, 'Gantt Task')
        r = client.get(f'/projects/{p.id}/gantt-data')
        assert r.status_code == 200
        data = r.get_json()
        assert len(data) == 1
        assert data[0]['name'] == 'Gantt Task'
        assert 'start' in data[0]
        assert 'end' in data[0]
        assert 'milestones' in data[0]


class TestEditProject:
    def test_get_edit_form(self, client, db):
        p = make_project()
        r = client.get(f'/projects/{p.id}/edit')
        assert r.status_code == 200
        assert b'Test Project' in r.data

    def test_update_project(self, client, db):
        p = make_project('Old Name')
        r = client.post(f'/projects/{p.id}/edit', data={
            'name': 'New Name',
            'status': 'completed',
            'start_date': '2025-01-01',
            'end_date': '2025-12-31',
        }, follow_redirects=True)
        assert r.status_code == 200
        db.session.refresh(p)
        assert p.name == 'New Name'
        assert p.status == 'completed'


class TestDeleteProject:
    def test_get_shows_confirmation(self, client, db):
        p = make_project('To Delete')
        r = client.get(f'/projects/{p.id}/delete')
        assert r.status_code == 200
        assert b'Delete Project' in r.data
        assert b'To Delete' in r.data

    def test_post_deletes_project(self, client, db):
        p = make_project('Gone Project')
        pid = p.id
        r = client.post(f'/projects/{pid}/delete', follow_redirects=True)
        assert r.status_code == 200
        assert Project.query.get(pid) is None

    def test_post_redirects_to_list(self, client, db):
        p = make_project()
        r = client.post(f'/projects/{p.id}/delete')
        assert r.status_code == 302
        assert '/projects' in r.headers['Location']

class TestExcelExport:
    def test_returns_xlsx(self, client, db):
        p = make_project('Export Me')
        r = client.get(f'/projects/{p.id}/export/excel')
        assert r.status_code == 200
        assert r.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def test_filename_contains_project_name(self, client, db):
        p = make_project('My Project')
        r = client.get(f'/projects/{p.id}/export/excel')
        disposition = r.headers.get('Content-Disposition', '')
        assert 'My_Project_gantt.xlsx' in disposition

    def test_spreadsheet_has_correct_headers(self, client, db):
        p = make_project()
        r = client.get(f'/projects/{p.id}/export/excel')
        wb = openpyxl.load_workbook(BytesIO(r.data))
        ws = wb['Gantt Data']
        headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
        assert headers == ['Task', 'Start Date', 'End Date', 'Status', 'Priority', 'Assignees', 'Tags']

    def test_spreadsheet_contains_task_data(self, client, db):
        p = make_project()
        make_task(p, 'Design Phase', status='in_progress', priority='high')
        r = client.get(f'/projects/{p.id}/export/excel')
        wb = openpyxl.load_workbook(BytesIO(r.data))
        ws = wb['Gantt Data']
        titles = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]
        assert 'Design Phase' in titles

    def test_spreadsheet_includes_milestones(self, client, db):
        p = make_project()
        t = make_task(p)
        make_milestone(t, 'Launch', date(2025, 6, 1))
        r = client.get(f'/projects/{p.id}/export/excel')
        wb = openpyxl.load_workbook(BytesIO(r.data))
        ws = wb['Gantt Data']
        # Milestone 1 header should be present
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert 'Milestone 1' in headers
        assert 'Milestone 1 Date' in headers

    def test_project_info_sheet_exists(self, client, db):
        p = make_project('Info Sheet Test')
        r = client.get(f'/projects/{p.id}/export/excel')
        wb = openpyxl.load_workbook(BytesIO(r.data))
        assert 'Project Info' in wb.sheetnames
        meta = wb['Project Info']
        names = [meta.cell(row=i, column=2).value for i in range(1, 4)]
        assert 'Info Sheet Test' in names

    def test_404_for_missing_project(self, client):
        r = client.get('/projects/99999/export/excel')
        assert r.status_code == 404


    def test_deletes_cascade_to_tasks(self, client, db):
        from models import Task
        p = make_project()
        t = make_task(p)
        tid = t.id
        client.post(f'/projects/{p.id}/delete')
        assert Task.query.get(tid) is None
